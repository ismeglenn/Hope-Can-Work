"""
update_recommendations.py

Fetches the latest recommendations from OWASP and updates recommendations.xlsx.
If recommendations.xlsx does not exist, it will be created automatically.

Run this script whenever you want to refresh the recommendations:
    python update_recommendations.py

Requirements:
    pip install requests beautifulsoup4 openpyxl
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

EXCEL_PATH = "Recommendations.xlsx"

OWASP_SOURCES = [
    {
        "url": "https://owasp.org/Top10/2025/A01_2025-Broken_Access_Control/",
        "violation_type": "BROKEN_ACCESS_CONTROL",
        "section": "How to Prevent",
    },
    {
        "url": "https://owasp.org/Top10/2025/A07_2025-Authentication_Failures/",
        "violation_type": "MISSING_AUTHENTICATION",
        "section": "How to Prevent",
    },
    {
        "url": "https://owasp.org/Top10/2025/A02_2025-Security_Misconfiguration/",
        "violation_type": "SECURITY_MISCONFIGURATION",
        "section": "How to Prevent",
    },
    {
        "url": "https://owasp.org/Top10/2025/A04_2025-Cryptographic_Failures/",
        "violation_type": "SENSITIVE_DATA_EXPOSURE",
        "section": "How to Prevent",
    },
]

HEADERS = ["violation_type", "keyword", "recommendation", "owasp_reference", "severity_hint", "link"]

HEADER_FONT       = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL       = PatternFill("solid", start_color="16213E")
HEADER_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT         = Font(name="Arial", size=9)
CELL_ALIGN        = Alignment(vertical="top", wrap_text=True)
ALT_FILL          = PatternFill("solid", start_color="EEF2FF")
WHITE_FILL        = PatternFill("solid", start_color="FFFFFF")
THIN              = Side(style="thin", color="CCCCCC")
BORDER            = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = [30, 15, 80, 45, 15, 50]  


def clean_text(text: str) -> str:
    """Replace special unicode characters that cannot be encoded in latin-1."""
    replacements = {
        '\u2014': '-',   # em dash —
        '\u2013': '-',   # en dash –
        '\u2018': "'",   # left single quote '
        '\u2019': "'",   # right single quote '
        '\u201c': '"',   # left double quote "
        '\u201d': '"',   # right double quote "
        '\u2026': '...', # ellipsis …
        '\u00a0': ' ',   # non-breaking space
        '\u2022': '-',   # bullet •
    }
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    return text

def create_empty_excel(path: str):
    """Create a fresh recommendations.xlsx with headers and instructions sheet."""
    print(f"  '{path}' not found — creating new file...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"

    # Header row
    ws.append(HEADERS)
    for col, _ in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # Instructions sheet
    ws2 = wb.create_sheet("How to Use")
    instructions = [
        ["Field",          "Description"],
        ["violation_type", "Matches violation_type in access_control_report.json"],
        ["keyword",        "Word matched against violated URL. Use 'default' as fallback."],
        ["recommendation", "Full recommendation text shown in PDF report"],
        ["owasp_reference","OWASP Top 10 reference this violation maps to"],
        ["severity_hint",  "Suggested severity: LOW, MEDIUM, HIGH, CRITICAL"],
        ["link", "Direct URL to the OWASP page this recommendation came from"],
        ["", ""],
        ["How it works",   ""],
        ["1.", "Python reads access_control_report.json and extracts violations"],
        ["2.", "For each violation, looks up violation_type in this Excel file"],
        ["3.", "Tries to match the violated URL against the keyword column"],
        ["4.", "If a keyword matches, that recommendation is used"],
        ["5.", "If no keyword matches, the 'default' row is used as fallback"],
        ["6.", "To add new violation types, add new rows to Recommendations sheet"],
    ]
    for i, row in enumerate(instructions, 1):
        for j, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=j)
            cell.value = val
            cell.font = Font(name="Arial", bold=(i == 1 or j == 1), size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 90

    wb.save(path)
    print(f"  Created '{path}' successfully.")


def fetch_owasp_recommendations(url: str, section: str = "How to Prevent") -> list:
    """Fetch the How to Prevent section from an OWASP Top 10 page."""
    print(f"  Fetching: {url}")
    headers = {"User-Agent": "Mozilla/5.0 (compatible; SecurityScanner/1.0)"}

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.encoding = 'utf-8'
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"  ERROR fetching {url}: {e}")
        return []

    soup = BeautifulSoup(response.text, "html.parser")

    target_heading = None
    for heading in soup.find_all(["h2", "h3"]):
        if section.lower() in heading.get_text(strip=True).lower():
            target_heading = heading
            break

    if not target_heading:
        print(f"  WARNING: Could not find section '{section}' on {url}")
        return []

    recommendations = []
    for sibling in target_heading.find_next_siblings():
        if sibling.name in ["h2", "h3"]:
            break
        if sibling.name in ["ul", "ol"]:
            for li in sibling.find_all("li"):
                text = clean_text(li.get_text(separator=" ", strip=True))
                if text:
                    recommendations.append(text)
        elif sibling.name == "p":
            text = clean_text(sibling.get_text(separator=" ", strip=True))
            if text:
                recommendations.append(text)

    print(f"  Found {len(recommendations)} recommendations")
    return recommendations


def build_keyword_map(recommendations: list, source_url: str) -> list:
    """Map each recommendation to a keyword based on its content."""
    keyword_hints = {
        "logout":   ["logout", "session invalidat", "sign out"],
        "admin":    ["admin", "privilege", "superuser", "root"],
        "delete":   ["delet", "remov"],
        "cart":     ["cart", "basket", "purchase"],
        "file":     ["file", "download", "upload", "attachment"],
        "password": ["password", "credential", "secret"],
        "role":     ["role", "permission", "access control list", "acl"],
        "api":      ["api", "endpoint", "service", "microservice"],
        "token":    ["token", "jwt", "bearer", "oauth"],
        "log":      ["log", "audit", "monitor", "alert"],
    }

    mapped = []
    used_keywords = set()

    for rec in recommendations:
        rec_lower = rec.lower()
        matched_keyword = "default"

        for keyword, hints in keyword_hints.items():
            if keyword in used_keywords:
                continue
            if any(hint in rec_lower for hint in hints):
                matched_keyword = keyword
                used_keywords.add(keyword)
                break

        # Return only keyword, recommendation, and link (title removed)
        mapped.append((matched_keyword, rec, source_url))

    # Ensure there's always a default row
    if not any(m[0] == "default" for m in mapped) and mapped:
        last = mapped[-1]
        mapped[-1] = ("default", last[1], last[2])

    return mapped


def update_excel(excel_path: str, violation_type: str, mapped_recommendations: list):
    """Remove old rows for this violation_type and insert fresh ones."""
    wb = load_workbook(excel_path)
    ws = wb["Recommendations"]

    # Remove old rows for this violation_type
    rows_to_delete = [
        row[0].row for row in ws.iter_rows(min_row=2)
        if row[0].value == violation_type
    ]
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)
    print(f"  Removed {len(rows_to_delete)} old rows for {violation_type}")

    owasp_ref = clean_text(
        f"OWASP - {violation_type.replace('_', ' ').title()} "
        f"(Updated {datetime.now().strftime('%Y-%m-%d')})"
    )

    for keyword, recommendation, link in mapped_recommendations:
        # Clean EVERY field that will go into Excel
        clean_keyword = clean_text(keyword)
        clean_recommendation = clean_text(recommendation)
        clean_link = clean_text(link)
        
        ws.append([violation_type, clean_keyword, clean_recommendation, owasp_ref, "MEDIUM", clean_link])

    # Append new rows
    owasp_ref = (
        f"OWASP - {violation_type.replace('_', ' ').title()} "
        f"(Updated {datetime.now().strftime('%Y-%m-%d')})"
    )
    for keyword, recommendation, link in mapped_recommendations:
        ws.append([violation_type, keyword, recommendation, owasp_ref, "MEDIUM", link])

    # Reapply formatting to all data rows
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for cell in row:
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = BORDER
            cell.fill = fill

    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 60

    wb.save(excel_path)
    print(f"  Saved {len(mapped_recommendations)} new rows for {violation_type}")


def main():
    print("=" * 60)
    print("OWASP Recommendations Updater")
    print("=" * 60)

    # Auto-create Excel if it doesn't exist
    if not os.path.exists(EXCEL_PATH):
        create_empty_excel(EXCEL_PATH)

    for source in OWASP_SOURCES:
        print(f"\nProcessing: {source['violation_type']}")
        print("-" * 40)

        recommendations = fetch_owasp_recommendations(source["url"], source["section"])

        if not recommendations:
            print(f"  Skipping {source['violation_type']} — no recommendations fetched.")
            continue

        mapped = build_keyword_map(recommendations, source["url"])
        update_excel(EXCEL_PATH, source["violation_type"], mapped)

    print("\n" + "=" * 60)
    print(f"Done! '{EXCEL_PATH}' is up to date with latest OWASP recommendations.")
    print("=" * 60)


if __name__ == "__main__":
    main()